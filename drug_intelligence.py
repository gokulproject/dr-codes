"""
checker.py — Oracle DB connection check logic.

Flow per customer/environment:
  1. Load customer from health_check.Customers (read-only)
  2. Load environments from health_check.Environments (read-only)
  3. Map DB server from health_check.Servers (ServerName='DB Server')
  4. Load Oracle DBs from health_check.OracleDatabase (filtered by db_type)
  5. Connect using oracledb — thick if oracle_dll_path set, else thin
  6. Classify any error: TCP_TIMEOUT | PASSWORD | UNKNOWN
  7. Write one row per Oracle DB to process_db_connection_result

No WMI. No connection timeout configured (OS/driver default applies).
"""
import logging
import os
import sys
import time
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from Common.crypto     import decrypt_password, is_encrypted
from Common.db_manager import DBManager
from config            import HC_DB, FMWCHECKS_DB

log = logging.getLogger("checker.checker")

ALL_ENV_TYPES = ["DEV", "VAL", "PRD"]
IST           = timedelta(hours=5, minutes=30)

# ── Error classification ──────────────────────────────────────────────────────
_TCP_KEYWORDS = [
    "ora-12170", "ora-12541", "ora-12543", "ora-12547",
    "tns-12535", "tns-12606", "tns-12609",
    "connection timed out", "timed out", "timeout",
    "connect timeout", "tcp", "network adapter",
    "no listener", "unable to connect to oracle",
]
_PWD_KEYWORDS = [
    "ora-28001", "ora-28002", "ora-28003", "ora-28000", "ora-28009",
    "ora-01017", "ora-01005",
    "invalid username/password", "logon denied",
    "password expired", "password has expired", "account is locked",
]

ERROR_TCP = "TCP_TIMEOUT"
ERROR_PWD = "PASSWORD"
ERROR_UNK = "UNKNOWN"


def classify_error(msg: str) -> str:
    if not msg:
        return ERROR_UNK
    lower = msg.lower()
    if any(k in lower for k in _TCP_KEYWORDS):
        return ERROR_TCP
    if any(k in lower for k in _PWD_KEYWORDS):
        return ERROR_PWD
    return ERROR_UNK


class Checker:

    def __init__(self, run_id: int, app_cfg=None):
        self.run_id  = run_id
        self.cfg     = app_cfg
        self._dll    = getattr(app_cfg, "oracle_dll_path",  "") if app_cfg else ""
        self._custs  = getattr(app_cfg, "config_customers", []) if app_cfg else []
        self._envs   = getattr(app_cfg, "config_envs",      []) if app_cfg else []
        self.db_types = getattr(app_cfg, "config_dbtype", ["FMW"]) if app_cfg else ["FMW"]
        self.db_label = getattr(app_cfg, "report_label", "/".join(self.db_types)) \
            if app_cfg else "/".join(self.db_types)
        self._thick  = False
        self._init_oracle_client()

    # ── Oracle client init ────────────────────────────────────────────────────

    def _init_oracle_client(self):
        try:
            import oracledb
            dll = self._dll
            if not dll or not dll.strip():
                log.warning("  Oracle DLL path not set in app_config — running in THIN mode")
                log.warning("  Set oracle_dll_path to the Instant Client folder to enable thick mode")
                return
            if not os.path.isdir(dll):
                log.warning(f"  Oracle DLL path not found: {dll} — running in THIN mode")
                return
            if not os.listdir(dll):
                log.warning(f"  Oracle DLL folder is empty: {dll} — running in THIN mode")
                return
            oracledb.init_oracle_client(lib_dir=dll)
            self._thick = True
            log.info(f"  Oracle THICK mode initialised  |  path = {dll}")
        except Exception as e:
            log.warning(f"  Oracle client init failed ({e}) — running in THIN mode")

    # ── Oracle DB connection check + FMW query (single connection) ───────────

    def _check_db(self, info: dict) -> tuple:
        # Opens ONE connection per DB entry — connectivity check + FMW query on same conn.
        # Returns: (connected, error_msg, error_type, duration_ms, fmw_rows, fmw_query_ran)
        # fmw_query_ran = True  → query executed successfully (rows may still be empty)
        # fmw_query_ran = False → query never ran (not FMW, failed conn, error, not configured)
        import oracledb

        host    = info["db_host"]
        port    = int(info.get("db_port", 1521))
        service = info["db_service"]
        user    = info["db_user"]
        raw_pwd = info["db_password"]
        name    = info["db_name"]
        cname   = info["customer_name"]
        env     = info["env_type"]
        dtype   = info.get("db_type", "").upper()

        try:
            pwd = decrypt_password(raw_pwd) if is_encrypted(raw_pwd) else raw_pwd
        except Exception as e:
            log.warning(f"    Password decrypt warning ({e}) — using raw value")
            pwd = raw_pwd

        dsn  = f"{host}:{port}/{service}"
        mode = "THICK" if self._thick else "THIN"
        log.info(f"    Connecting  {cname} / {env} / {name}  [{dtype}]  {dsn}  ({mode})")

        start         = time.time()
        conn          = None
        fmw_rows      = []
        fmw_query_ran = False  # stays False unless query runs without exception

        try:
            conn = oracledb.connect(user=user, password=pwd, dsn=dsn)
            ms   = int((time.time() - start) * 1000)
            log.info(f"    CONNECTED   {cname} / {env} / {name}  ({ms} ms)")

            # FMW query runs on the same open connection — only for FMW db_type
            if dtype == "FMW":
                fmw_query = getattr(self.cfg, "fmw_query", "") if self.cfg else ""
                if not fmw_query:
                    # fmw_query_ran stays False — query not configured
                    log.warning("    FMW query not configured in app_config — skipping")
                else:
                    try:
                        log.info(f"    FMW Query  {cname} / {env} / {name}  —  executing OAM user query")
                        cursor = conn.cursor()
                        cursor.execute(fmw_query)
                        cols = [desc[0].upper() for desc in cursor.description]
                        for row in cursor.fetchall():
                            row_dict = {}
                            for col, val in zip(cols, row):
                                if hasattr(val, "strftime"):
                                    row_dict[col] = val.strftime("%d-%b-%Y")
                                else:
                                    row_dict[col] = str(val) if val is not None else ""
                            fmw_rows.append(row_dict)
                        fmw_query_ran = True  # query ran — zero rows is still a valid result
                        log.info(f"    FMW Query  {cname} / {env} / {name}  —  {len(fmw_rows)} OAM account(s) returned")
                    except Exception as qe:
                        fmw_query_ran = False  # query threw an error — treat as not ran
                        log.error(f"    FMW Query  {cname} / {env} / {name}  —  query failed: {qe}")
                        fmw_rows = []

            return True, None, "", ms, fmw_rows, fmw_query_ran

        except Exception as e:
            ms       = int((time.time() - start) * 1000)
            err      = str(e)
            err_type = classify_error(err)
            log.error(f"    FAILED      {cname} / {env} / {name}  ({ms} ms)  [{err_type}]")
            log.error(f"                {err}")
            # Connection failed — fmw_query_ran stays False
            return False, err, err_type, ms, [], False

        finally:
            # Single close — connection opened exactly once above
            if conn:
                try:   conn.close()
                except Exception: pass

    # ── health_check reads (READ ONLY) ────────────────────────────────────────

    def _customers(self, db):
        if self._custs:
            ph = ",".join(["%s"] * len(self._custs))
            rows = db.fetch_all(
                f"SELECT CustomerID AS customer_id, CustomerName AS customer_name "
                f"FROM Customers WHERE CustomerName IN ({ph}) AND Status=1 "
                f"ORDER BY CustomerName",
                tuple(self._custs))
        else:
            rows = db.fetch_all(
                "SELECT CustomerID AS customer_id, CustomerName AS customer_name "
                "FROM Customers WHERE Status=1 ORDER BY CustomerName")
        log.info(f"  Customers loaded  :  {len(rows)} found"
                 + (f"  (filter: {self._custs})" if self._custs else "  (all active)"))
        return rows

    def _envs_for(self, db, cid):
        scope = self._envs or ALL_ENV_TYPES
        ph    = ",".join(["%s"] * len(scope))
        return db.fetch_all(
            f"SELECT EnvID AS env_id, ServerType AS env_type "
            f"FROM Environments "
            f"WHERE CustomerID=%s AND ServerType IN ({ph}) AND Status=1 "
            f"ORDER BY FIELD(ServerType,'DEV','VAL','PRD')",
            tuple([cid] + scope))

    def _server_for(self, db, env_id):
        return db.fetch_one(
            "SELECT ServerID AS server_id, ServerName AS server_name, Host AS host "
            "FROM Servers "
            "WHERE EnvID=%s AND ServerName='DB Server' AND Status=1 LIMIT 1",
            (env_id,))

    def _oracle_dbs(self, db, server_id):
        ph = ",".join(["%s"] * len(self.db_types))
        return db.fetch_all(
            f"SELECT dbid AS db_id, OrdName AS db_name, OrdName AS db_service, "
            f"Port AS db_port, User AS db_user, Password AS db_password, "
            f"DatabaseType AS db_type "
            f"FROM OracleDatabase "
            f"WHERE ServerID=%s AND DatabaseType IN ({ph}) AND Status=1",
            tuple([server_id] + self.db_types))

    # ── Result helpers ────────────────────────────────────────────────────────

    def _row(self, cid, cname, env_type,
             server_id=0, server_name="", server_host="",
             db_name="", db_type=None, db_host="", db_port=0, db_service="",
             status="SKIPPED", error="", error_type="", ms=0) -> dict:
        return {
            "run_id": self.run_id, "customer_id": cid,
            "customer_name": cname, "env_type": env_type,
            "server_id": server_id, "server_name": server_name,
            "server_host": server_host,
            "db_name": db_name,
            "db_type": db_type if db_type is not None else "/".join(self.db_types),
            "db_host": db_host, "db_port": db_port, "db_service": db_service,
            "connection_status": status, "error_message": error,
            "error_type": error_type, "duration_ms": ms,
            "checked_at": datetime.utcnow(),
            "fmw_query_rows": [],   # skipped rows — no query ran
            "fmw_query_ran":  False,
        }

    def _save(self, result: dict):
        try:
            with DBManager(cfg=FMWCHECKS_DB) as db:
                db.execute(
                    "INSERT INTO process_db_connection_result "
                    "(run_id,customer_id,customer_name,env_type,"
                    "server_id,server_name,server_host,"
                    "db_name,db_type,db_host,db_port,db_service,"
                    "connection_status,error_message,error_type,"
                    "duration_ms,checked_at) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (result["run_id"],
                     result["customer_id"], result["customer_name"], result["env_type"],
                     result["server_id"],   result["server_name"],   result["server_host"],
                     result["db_name"],     result["db_type"],       result["db_host"],
                     result["db_port"],     result["db_service"],
                     result["connection_status"], result["error_message"],
                     result["error_type"],  result["duration_ms"],   result["checked_at"]))
        except Exception as e:
            log.error(f"  DB write failed: {e}")

    def _save_fmw_results(self, run_id: int, result: dict, fmw_rows: list):
        # Writes each OAM account row to fmw_query_result — called only when query ran successfully
        customer_name = result.get("customer_name", "")
        env_type      = result.get("env_type",      "")
        server_host   = result.get("server_host",   "")
        db_name       = result.get("db_name",       "")
        db_type       = result.get("db_type",       "")
        conn_ref      = f"{server_host} / {db_name} / {db_type}"

        if not fmw_rows:
            # Query ran but returned zero OAM accounts — write one blank row to capture the event
            try:
                with DBManager(cfg=FMWCHECKS_DB) as db:
                    db.execute(
                        "INSERT INTO fmw_query_result "
                        "(run_id, customer_name, env_type, server_db_info, "
                        " username, account_status, expiry_date, "
                        " action_required, reason, queried_at) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
                        (run_id, customer_name, env_type, conn_ref,
                         "", "", "", "No", "",))
            except Exception as e:
                log.error(f"  fmw_query_result write failed (no-rows): {e}")
            return

        for frow in fmw_rows:
            expiry          = frow.get("EXPIRY_DATE", "") or ""
            expiry_set      = expiry.strip().upper() not in ("", "NULL", "NONE")
            action_required = "Yes" if expiry_set else "No"
            reason          = expiry.strip() if expiry_set else ""
            try:
                with DBManager(cfg=FMWCHECKS_DB) as db:
                    db.execute(
                        "INSERT INTO fmw_query_result "
                        "(run_id, customer_name, env_type, server_db_info, "
                        " username, account_status, expiry_date, "
                        " action_required, reason, queried_at) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
                        (run_id, customer_name, env_type, conn_ref,
                         frow.get("USERNAME",       ""),
                         frow.get("ACCOUNT_STATUS", ""),
                         expiry,
                         action_required,
                         reason))
            except Exception as e:
                log.error(f"  fmw_query_result write failed: {e}")

    def _log_step(self, no, name, status, msg="", cust="", env=""):
        try:
            with DBManager(cfg=FMWCHECKS_DB) as db:
                db.execute(
                    "INSERT INTO process_run_log "
                    "(run_id,step_no,step_name,status,message,customer,env_type,logged_at) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())",
                    (self.run_id, no, name, status,
                     (msg  or "")[:2000],
                     (cust or "")[:255],
                     (env  or "")[:50]))
        except Exception as e:
            log.warning(f"  Step log write failed: {e}")

    # ── Main loop ─────────────────────────────────────────────────────────────

    def run_all(self) -> list:
        results   = []
        env_scope = self._envs or ALL_ENV_TYPES

        log.info(f"[Step 3]  Loading customers  |  db_types={self.db_types}  |  envs={env_scope}")
        self._log_step(3, "IDENTIFY_CUSTOMERS", "STARTED",
                       f"customer_filter={'ALL' if not self._custs else self._custs}  |  "
                       f"env_filter={'ALL' if not self._envs else self._envs}  |  "
                       f"db_type={self.db_types}")

        try:
            with DBManager(cfg=HC_DB) as db:
                customers = self._customers(db)
        except Exception as e:
            log.error(f"  Cannot load customers: {e}")
            self._log_step(3, "IDENTIFY_CUSTOMERS", "FAILED", str(e))
            return []

        if not customers:
            log.warning("  No active customers found")
            self._log_step(3, "IDENTIFY_CUSTOMERS", "SKIPPED", "No active customers")
            return []

        self._log_step(3, "IDENTIFY_CUSTOMERS", "COMPLETED",
                       f"{len(customers)} customer(s) found")

        for cust in customers:
            cid   = cust["customer_id"]
            cname = cust["customer_name"]

            log.info("")
            log.info(f"  ── Customer: {cname}  (ID={cid})")

            # Load environments
            try:
                with DBManager(cfg=HC_DB) as db:
                    envs = self._envs_for(db, cid)
            except Exception as e:
                log.error(f"    Cannot load environments: {e}")
                self._log_step(4, "IDENTIFY_ENVIRONMENTS", "FAILED", str(e), cname)
                continue

            if not envs:
                msg = f"No active environments found (filter={env_scope})"
                log.warning(f"    {msg}")
                self._log_step(4, "IDENTIFY_ENVIRONMENTS", "SKIPPED", msg, cname)
                r = self._row(cid, cname, ",".join(env_scope),
                              error=msg, status="SKIPPED")
                self._save(r); results.append(r)
                continue

            log.info(f"    Environments : {[e['env_type'] for e in envs]}")
            self._log_step(4, "IDENTIFY_ENVIRONMENTS", "COMPLETED",
                           f"found: {[e['env_type'] for e in envs]}", cname)

            for env in envs:
                env_id   = env["env_id"]
                env_type = env["env_type"]

                log.info(f"    ── Env: {env_type}  (EnvID={env_id})")

                # Map DB server
                try:
                    with DBManager(cfg=HC_DB) as db:
                        server = self._server_for(db, env_id)
                except Exception as e:
                    log.error(f"      Server lookup failed: {e}")
                    self._log_step(5, "MAP_DB_SERVER", "FAILED", str(e), cname, env_type)
                    r = self._row(cid, cname, env_type,
                                  error=f"Server lookup failed: {e}", status="SKIPPED")
                    self._save(r); results.append(r)
                    continue

                if not server:
                    msg = "No active 'DB Server' row in health_check.Servers"
                    log.warning(f"      {msg}")
                    self._log_step(5, "MAP_DB_SERVER", "SKIPPED", msg, cname, env_type)
                    r = self._row(cid, cname, env_type, error=msg, status="SKIPPED")
                    self._save(r); results.append(r)
                    continue

                server_id   = server["server_id"]
                server_name = server["server_name"]
                server_host = server["host"]
                log.info(f"      DB Server  :  {server_name}  |  host={server_host}  (ID={server_id})")
                self._log_step(5, "MAP_DB_SERVER", "COMPLETED",
                               f"name={server_name} id={server_id} host={server_host}",
                               cname, env_type)

                # Load Oracle DBs
                try:
                    with DBManager(cfg=HC_DB) as db:
                        dbs = self._oracle_dbs(db, server_id)
                except Exception as e:
                    log.error(f"      Oracle DB lookup failed: {e}")
                    self._log_step(6, "LOAD_DB_TYPES", "FAILED", str(e), cname, env_type)
                    r = self._row(cid, cname, env_type,
                                  server_id=server_id, server_name=server_name,
                                  server_host=server_host, db_host=server_host,
                                  error=f"DB lookup failed: {e}", status="SKIPPED")
                    self._save(r); results.append(r)
                    continue

                if not dbs:
                    msg = f"No active Oracle DB found for db_type={self.db_types}"
                    log.warning(f"      {msg}")
                    self._log_step(6, "LOAD_DB_TYPES", "SKIPPED", msg, cname, env_type)
                    r = self._row(cid, cname, env_type,
                                  server_id=server_id, server_name=server_name,
                                  server_host=server_host, db_host=server_host,
                                  error=msg, status="SKIPPED")
                    self._save(r); results.append(r)
                    continue

                log.info(f"      Oracle DBs : {len(dbs)} found  (type={self.db_types})")
                self._log_step(6, "LOAD_DB_TYPES", "COMPLETED",
                               f"{len(dbs)} DB(s)", cname, env_type)

                for tdb in dbs:
                    info = {
                        "customer_id":  cid,
                        "customer_name": cname,
                        "env_type":     env_type,
                        "db_name":      tdb["db_name"],
                        "db_type":      tdb.get("db_type", ""),
                        "db_host":      server_host,
                        "db_port":      tdb.get("db_port", 1521),
                        "db_service":   tdb["db_service"],
                        "db_user":      tdb["db_user"],
                        "db_password":  tdb["db_password"],
                    }

                    # Unpack 6 values — fmw_query_ran distinguishes "ran ok" from "error/not ran"
                    ok, err, err_type, ms, fmw_query_rows, fmw_query_ran = self._check_db(info)
                    status = "COMPLETED" if ok else "FAILED"

                    r = {
                        "run_id":            self.run_id,
                        "customer_id":       cid,
                        "customer_name":     cname,
                        "env_type":          env_type,
                        "server_id":         server_id,
                        "server_name":       server_name,
                        "server_host":       server_host,
                        "db_name":           tdb["db_name"],
                        "db_type":           tdb.get("db_type", ""),
                        "db_host":           server_host,
                        "db_port":           tdb.get("db_port", 1521),
                        "db_service":        tdb["db_service"],
                        "connection_status": status,
                        "error_message":     err or "",
                        "error_type":        err_type if not ok else "",
                        "duration_ms":       ms,
                        "checked_at":        datetime.utcnow(),
                        "fmw_query_rows":    fmw_query_rows,
                        "fmw_query_ran":     fmw_query_ran,  # True = query ran ok, False = did not run
                    }
                    self._save(r)

                    # Save to fmw_query_result only when query actually ran successfully
                    if tdb.get("db_type", "").upper() == "FMW" and fmw_query_ran:
                        self._save_fmw_results(self.run_id, r, fmw_query_rows)

                    results.append(r)
                    self._log_step(7, "DB_CONNECTION_CHECK", status,
                                   f"DB={tdb['db_name']}  |  {ms}ms"
                                   + (f"  |  [{err_type}] {err}" if err else ""),
                                   cname, env_type)

        # Summary
        s = sum(1 for r in results if r["connection_status"] == "COMPLETED")
        f = sum(1 for r in results if r["connection_status"] == "FAILED")
        k = sum(1 for r in results if r["connection_status"] == "SKIPPED")
        log.info("")
        log.info(f"[Step 7]  Check complete  —  {len(results)} total  |  "
                 f"connected={s}  |  not_connected={f}  |  skipped={k}")
        return results

----------------------------

"""
ExcelCreation/excel_manager.py
==============================
Generate Excel report for Oracle DB Connection Check.

Sheet 1 — DB Connection Check (16 columns):
  # | Customer | Environment | Server Name | Server ID | Server Host |
  DB Name | DB Type | DSN | DB Connection Status |
  Duration (ms) | Error | Error Type | Checked At (IST) |
  Expired (Action Required) | Reason (if yes)

Dynamic sheets — one per Customer + Environment (FMW db_type only):
  Sheet name : "CustomerName - ENV"
  Row 2      : FMW query read from app_config
  Columns    : Server Host / DB Name / DB Type | USERNAME | ACCOUNT_STATUS |
               EXPIRY_DATE | Action Required

Expiry / Action Required rules:
  FMW + CONNECTED + query ran   + date found   → Yes / Yes | Reason: date
  FMW + CONNECTED + query ran   + no date      → No  / No  | Reason: blank
  FMW + CONNECTED + query error / not configured → blank   | Reason: blank
  FMW + NOT CONNECTED or SKIPPED               → blank     | Reason: blank
  Non-FMW db_type (any status)                 → blank     | Reason: blank
"""
import logging
import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log      = logging.getLogger("checker.excel")
IST      = timedelta(hours=5, minutes=30)
LAST_COL = 16   # 14 original + 2 expiry columns

# ── Palette ───────────────────────────────────────────────────────────────────
HDR_DARK   = "1565C0"
HDR_MED    = "0D47A1"
HDR_GREEN  = "1B5E20"
SUMM_BG    = "E3F2FD"
GREEN_BG   = "E8F5E9";  GREEN_BG2 = "F1F8E9"
RED_BG     = "FFEBEE";  RED_BG2   = "FFF5F5"
AMBER_BG   = "FFF8E1";  AMBER_BG2 = "FFFDE7"
C_GREEN    = "2E7D32"
C_RED      = "C62828"
C_AMBER    = "E65100"
C_BLUE     = "0D47A1"
C_WHITE    = "FFFFFF"
C_DARK_RED = "B71C1C"
C_DARK_GRN = "1B5E20"

STATUS_LABEL = {
    "COMPLETED": "DB CONNECTED",
    "FAILED":    "DB NOT CONNECTED",
    "SKIPPED":   "SKIPPED",
}
STATUS_COLOR = {
    "COMPLETED": C_GREEN,
    "FAILED":    C_RED,
    "SKIPPED":   C_AMBER,
}
ERR_LABEL = {
    "TCP_TIMEOUT": "TCP Connection Timeout",
    "PASSWORD":    "Password Expired / Invalid Username or Password",
    "UNKNOWN":     "Unknown / Other Error",
    "":            "",
}
ERR_COLOR = {"TCP_TIMEOUT": C_RED, "PASSWORD": C_AMBER, "UNKNOWN": C_BLUE, "": ""}

_DEFAULT_FMW_QUERY = (
    "SELECT USERNAME, ACCOUNT_STATUS, EXPIRY_DATE "
    "FROM DBA_USERS WHERE USERNAME LIKE '%OAM%'"
)


def _border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _ist(dt) -> str:
    return (dt + IST).strftime("%d-%b-%Y %I:%M:%S %p IST") if dt else ""


def _col(n):
    return get_column_letter(n)


# ── Expiry helpers ────────────────────────────────────────────────────────────

def _is_expiry_set(val) -> bool:
    # True only when EXPIRY_DATE contains an actual date — not null/none/empty
    if val is None:
        return False
    return str(val).strip().upper() not in ("", "NULL", "NONE")


def _compute_expiry_status(fmw_query_rows: list) -> tuple:
    # Called only for rows where query ran successfully — returns (expired, action, reason)
    if not fmw_query_rows:
        return "No", "No", ""  # query ran but zero OAM accounts found
    for row in fmw_query_rows:
        expiry = row.get("EXPIRY_DATE", "")
        if _is_expiry_set(expiry):
            return "Yes", "Yes", str(expiry).strip()  # first account with a date wins
    return "No", "No", ""  # all accounts have empty/null expiry


def _build_expiry_map(results: list) -> dict:
    # Builds (customer_name, env_type) → (expired, action, reason)
    # Only includes FMW entries where connection COMPLETED and query actually ran
    env_rows: dict = {}
    for r in results:
        if r.get("db_type", "").upper() != "FMW":
            continue
        if r.get("connection_status") != "COMPLETED":
            continue  # failed/skipped — expiry columns stay blank
        if not r.get("fmw_query_ran", False):
            continue  # connected but query errored or not configured — stay blank
        key = (r.get("customer_name", ""), r.get("env_type", ""))
        env_rows.setdefault(key, []).extend(r.get("fmw_query_rows", []))
    return {key: _compute_expiry_status(rows) for key, rows in env_rows.items()}


# ── Sheet 1: DB Connection Check ─────────────────────────────────────────────

def _write_connection_sheet(ws, results, run_id, timestamp, db_label, expiry_map):
    total   = len(results)
    success = sum(1 for r in results if r.get("connection_status") == "COMPLETED")
    failed  = sum(1 for r in results if r.get("connection_status") == "FAILED")
    skipped = sum(1 for r in results if r.get("connection_status") == "SKIPPED")
    tcp_err = sum(1 for r in results if r.get("error_type") == "TCP_TIMEOUT")
    pwd_err = sum(1 for r in results if r.get("error_type") == "PASSWORD")
    unk_err = sum(1 for r in results
                  if r.get("error_type") == "UNKNOWN"
                  and r.get("connection_status") == "FAILED")

    if total == 0:
        summary, summ_color = "No data processed", C_RED
    elif failed == 0 and skipped == 0:
        summary, summ_color = "ALL CONNECTED  ✅", C_GREEN
    elif success == 0:
        summary, summ_color = (
            f"NO SUCCESSFUL CONNECTIONS  ❌   "
            f"Not Connected: {failed}  |  Skipped: {skipped}"), C_RED
    else:
        summary, summ_color = (
            f"PARTIAL  ⚠    Connected: {success}  |  "
            f"Not Connected: {failed}  |  Skipped: {skipped}"), C_AMBER

    last_letter = _col(LAST_COL)

    # Row 1 — Title
    ws.merge_cells(f"A1:{last_letter}1")
    c = ws["A1"]
    c.value     = f"{db_label} Oracle DB Connection Check  |  Run #{run_id}  |  {timestamp}"
    c.font      = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=HDR_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2 — Summary bar
    ws.merge_cells(f"A2:{last_letter}2")
    c = ws["A2"]
    err_part = (
        f"  |  Errors — TCP: {tcp_err}  |  Password: {pwd_err}  |  Unknown: {unk_err}"
        if failed else "")
    c.value = (
        f"Total: {total}   ✅ Connected: {success}   ❌ Not Connected: {failed}"
        f"   ⚠ Skipped: {skipped}{err_part}   —   {summary}")
    c.font      = Font(name="Calibri", size=10, bold=True, color=summ_color)
    c.fill      = PatternFill("solid", fgColor=SUMM_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row 3 — Headers
    headers = [
        "#", "Customer", "Environment", "Server Name", "Server ID",
        "Server Host", "DB Name", "DB Type", "DSN (Host:Port/Service)",
        "DB Connection Status", "Duration (ms)", "Error", "Error Type",
        "Checked At (IST)",
        "Expired\n(Action Required)", "Reason\n(if yes)",
    ]
    widths = [4, 22, 13, 18, 10, 22, 20, 11, 32, 20, 13, 45, 35, 24, 18, 45]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c           = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=HDR_MED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[_col(col)].width = w
    ws.row_dimensions[3].height = 30

    ws.auto_filter.ref = f"A3:{last_letter}{3 + max(total, 1)}"

    for idx, r in enumerate(results, 1):
        row      = idx + 3
        stat     = r.get("connection_status", "SKIPPED")
        err_type = r.get("error_type", "") or ""

        bg = (GREEN_BG if idx % 2 else GREEN_BG2) if stat == "COMPLETED" else \
             (RED_BG   if idx % 2 else RED_BG2)   if stat == "FAILED"    else \
             (AMBER_BG if idx % 2 else AMBER_BG2)

        h   = r.get("db_host",    "") or ""
        p   = r.get("db_port",    1521) or 1521
        svc = r.get("db_service", "") or ""
        dsn = f"{h}:{p}/{svc}" if h else ""

        # Expiry: FMW + COMPLETED + query ran → compute value; everything else → blank
        db_type_val = r.get("db_type", "").upper()
        if db_type_val == "FMW" and stat == "COMPLETED" and r.get("fmw_query_ran", False):
            key = (r.get("customer_name", ""), r.get("env_type", ""))
            expired_str, action_str, reason_str = expiry_map.get(key, ("No", "No", ""))
            expiry_display = f"{expired_str} / {action_str}"
        else:
            # blank for: failed, skipped, non-FMW, query errored, query not configured
            expiry_display = ""
            reason_str     = ""

        values = [
            idx,
            r.get("customer_name",  "") or "",
            r.get("env_type",       "") or "",
            r.get("server_name",    "") or "",
            r.get("server_id",      "") or "",
            r.get("server_host",    "") or "",
            r.get("db_name",        "") or "—",
            r.get("db_type",        "") or "",
            dsn,
            STATUS_LABEL.get(stat, stat),
            r.get("duration_ms",    0),
            r.get("error_message",  "") or "",
            ERR_LABEL.get(err_type, err_type),
            _ist(r.get("checked_at")),
            expiry_display,
            reason_str,
        ]

        for col, val in enumerate(values, 1):
            c           = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 10:  # DB Connection Status — bold coloured
                c.font      = Font(name="Calibri", size=10, bold=True,
                                   color=STATUS_COLOR.get(stat, C_RED))
                c.alignment = Alignment(horizontal="center", vertical="center")

            if col == 11:  # Duration — right-align
                c.alignment = Alignment(horizontal="right", vertical="center")

            if col == 12 and val:  # Error — italic
                c.font = Font(name="Calibri", size=9, italic=True,
                              color=STATUS_COLOR.get(stat, C_RED))

            if col == 13 and val:  # Error Type — bold coloured
                c.font = Font(name="Calibri", size=9, bold=True,
                              color=ERR_COLOR.get(err_type, C_BLUE))

            if col == 15 and expiry_display:  # Expired/Action — only when value present
                c.font = Font(name="Calibri", size=10, bold=True,
                              color=C_DARK_RED if expiry_display.startswith("Yes") else C_DARK_GRN)
                c.alignment = Alignment(horizontal="center", vertical="center")

            if col == 16 and reason_str:  # Reason — italic red only when date present
                c.font = Font(name="Calibri", size=9, italic=True, color=C_DARK_RED)

        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"


# ── Dynamic per-Customer/Env sheets (FMW only) ───────────────────────────────

def _safe_sheet_name(name: str) -> str:
    # Sanitise to valid Excel sheet name — max 31 chars, no illegal chars
    for ch in r"\/:*?[]":
        name = name.replace(ch, " ")
    return name[:31].strip()


def _write_fmw_env_sheet(ws, cname: str, env_type: str,
                         env_results: list, run_id, timestamp, fmw_query: str):
    # One sheet per Customer+Env — FMW db_type only
    # Row 2 shows the live query from app_config
    # Action Required: blank when not connected/skipped or query errored; Yes/No when query ran
    SHEET_COLS  = 5
    last_letter = _col(SHEET_COLS)

    # Row 1 — Title with customer, env, run info
    ws.merge_cells(f"A1:{last_letter}1")
    c = ws["A1"]
    c.value     = (f"{cname}  |  {env_type}  |  "
                   f"FMW OAM Account Expiry  |  Run #{run_id}  |  {timestamp}")
    c.font      = Font(name="Calibri", size=12, bold=True, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=HDR_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2 — FMW query read live from app_config
    ws.merge_cells(f"A2:{last_letter}2")
    c = ws["A2"]
    c.value     = f"FMW Query : {fmw_query}"
    c.font      = Font(name="Calibri", size=9, italic=True, color="37474F")
    c.fill      = PatternFill("solid", fgColor="E8F5E9")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Row 3 — Headers
    headers = [
        "Server Host / DB Name / DB Type",
        "USERNAME", "ACCOUNT_STATUS", "EXPIRY_DATE", "Action Required",
    ]
    widths = [40, 26, 20, 22, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c           = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=HDR_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[_col(col)].width = w
    ws.row_dimensions[3].height = 22

    data_rows = []
    for r in env_results:
        conn_info = (f"{r.get('server_host', '') or ''} / "
                     f"{r.get('db_name',     '') or ''} / "
                     f"{r.get('db_type',     '') or ''}")
        stat          = r.get("connection_status", "FAILED")
        fmw_rows      = r.get("fmw_query_rows", [])
        fmw_query_ran = r.get("fmw_query_ran",  False)

        if stat != "COMPLETED":
            # Not connected or skipped — query never ran — all OAM columns blank
            data_rows.append({
                "conn_info": conn_info, "USERNAME": "",
                "ACCOUNT_STATUS": "", "EXPIRY_DATE": "", "action": "",
            })
        elif not fmw_query_ran:
            # Connected but query errored or not configured — leave action blank
            data_rows.append({
                "conn_info": conn_info, "USERNAME": "",
                "ACCOUNT_STATUS": "", "EXPIRY_DATE": "", "action": "",
            })
        elif fmw_rows:
            # Query ran and returned OAM accounts — one row per account
            for frow in fmw_rows:
                expiry = frow.get("EXPIRY_DATE", "") or ""
                data_rows.append({
                    "conn_info":      conn_info,
                    "USERNAME":       frow.get("USERNAME",       ""),
                    "ACCOUNT_STATUS": frow.get("ACCOUNT_STATUS", ""),
                    "EXPIRY_DATE":    expiry,
                    "action":         "Yes" if _is_expiry_set(expiry) else "No",
                })
        else:
            # Query ran successfully but zero OAM accounts returned
            data_rows.append({
                "conn_info": conn_info, "USERNAME": "",
                "ACCOUNT_STATUS": "", "EXPIRY_DATE": "", "action": "No",
            })

    ws.auto_filter.ref = f"A3:{last_letter}{3 + max(len(data_rows), 1)}"

    for idx, dr in enumerate(data_rows, 1):
        row      = idx + 3
        bg       = "E8F5E9" if idx % 2 else "F1F8E9"
        expiry   = dr["EXPIRY_DATE"]
        has_date = _is_expiry_set(expiry)
        action   = dr["action"]

        for col, val in enumerate(
            [dr["conn_info"], dr["USERNAME"], dr["ACCOUNT_STATUS"], expiry, action], 1
        ):
            c           = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 4 and has_date:  # EXPIRY_DATE — red bold when real date
                c.font      = Font(name="Calibri", size=10, bold=True, color=C_DARK_RED)
                c.alignment = Alignment(horizontal="center", vertical="center")

            if col == 3 and val:  # ACCOUNT_STATUS — red if expired/locked, green otherwise
                up = str(val).upper()
                c.font = Font(name="Calibri", size=10, bold=True, color=C_RED) \
                    if ("EXPIRED" in up or "LOCKED" in up) \
                    else Font(name="Calibri", size=10, color=C_DARK_GRN)

            if col == 5 and action:  # Action Required — only style when not blank
                c.font = Font(name="Calibri", size=10, bold=True,
                              color=C_DARK_RED if action == "Yes" else C_DARK_GRN)
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"

    if not data_rows:
        ws.merge_cells(f"A4:{last_letter}4")
        c = ws["A4"]
        c.value     = f"No FMW OAM accounts found for {cname} / {env_type}."
        c.font      = Font(name="Calibri", size=10, italic=True, color=C_AMBER)
        c.alignment = Alignment(horizontal="center", vertical="center")


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(results, run_id, timestamp, excel_path,
                   db_label="FMW", fmw_query: str = "") -> str:
    # Generates report Excel — Sheet 1 (all DBs) + one sheet per FMW Customer/Env
    os.makedirs(excel_path, exist_ok=True)

    label    = (db_label or "FMW").upper().replace("/", "_").replace(" ", "_")
    date_str = datetime.now().strftime("%d-%m-%y")
    filename = f"{label}_CONNECTION_CHECKS_{date_str}-{run_id}.xlsx"
    filepath = os.path.join(excel_path, filename)

    query_text = fmw_query.strip() if fmw_query else _DEFAULT_FMW_QUERY

    # Expiry map — only from FMW COMPLETED entries where query actually ran
    expiry_map = _build_expiry_map(results)

    wb = Workbook()

    # Sheet 1 — DB Connection Check (all entries)
    ws1 = wb.active
    ws1.title = "DB Connection Check"
    _write_connection_sheet(ws1, results, run_id, timestamp, db_label, expiry_map)

    # Dynamic sheets — one per unique Customer + Environment for FMW db_type only
    env_groups: dict = {}
    for r in results:
        if r.get("db_type", "").upper() != "FMW":
            continue
        key = (r.get("customer_name", "") or "", r.get("env_type", "") or "")
        env_groups.setdefault(key, []).append(r)

    used_names: dict = {}
    for (cname, env_type), env_results in env_groups.items():
        safe_name = _safe_sheet_name(f"{cname} - {env_type}")
        if safe_name in used_names:
            used_names[safe_name] += 1
            safe_name = _safe_sheet_name(f"{safe_name} ({used_names[safe_name]})")
        else:
            used_names[safe_name] = 1
        ws = wb.create_sheet(title=safe_name)
        _write_fmw_env_sheet(ws, cname, env_type, env_results,
                             run_id, timestamp, query_text)
        log.info(f"  FMW sheet   : '{safe_name}'  ({len(env_results)} DB entry(s))")

    wb.save(filepath)
    log.info(f"  Excel saved : {filepath}")
    return filepath
