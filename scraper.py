"""
Excel creation and manipulation module
Uses ExcelHandler class and openpyxl for Excel operations
"""

import os
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from logger import get_logger

# Import from ExcelHandler if available
try:
    from ExcelHandler import ExcelHandler
except ImportError:
    ExcelHandler = None


class ExcelManager:
    """Excel file manager for reading and writing operations"""
    
    def __init__(self):
        """Initialize Excel manager"""
        self.logger = get_logger()
        self.excel_handler = ExcelHandler() if ExcelHandler else None
    
    def read_excel_with_clean_columns(
        self,
        file_path: str,
        sheet_name: str,
        column_names: List[str],
        start_row: int = 0
    ) -> List[List[Any]]:
        """
        Read Excel file and return clean data
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to read
            column_names: List of column names
            start_row: Starting row index
        
        Returns:
            List of row data
        """
        try:
            self.logger.log_function_start(
                "read_excel_with_clean_columns",
                file=file_path,
                sheet=sheet_name
            )
            
            # Use ExcelHandler if available
            if self.excel_handler:
                data = self.excel_handler.read_excel_with_clean_columns(
                    file_path, sheet_name, column_names, start_row
                )
                self.logger.log_function_end("read_excel_with_clean_columns")
                return data
            
            # Fallback to openpyxl
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]
            
            # Find header row
            header_row = None
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if any(col in str(cell).strip() if cell else '' 
                      for cell in row for col in column_names):
                    header_row = row_idx
                    break
            
            if header_row is None:
                raise ValueError(f"Header row not found in sheet {sheet_name}")
            
            # Get column indices
            header = list(sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True
            ))[0]
            
            col_indices = {}
            for col_name in column_names:
                for idx, cell in enumerate(header):
                    if cell and col_name.lower() in str(cell).lower().strip():
                        col_indices[col_name] = idx
                        break
            
            # Read data
            data = []
            start_data_row = header_row + 1 + start_row
            
            for row_idx, row in enumerate(
                sheet.iter_rows(
                    min_row=start_data_row,
                    values_only=True
                ),
                start=start_data_row
            ):
                row_data = [row_idx]
                for col_name in column_names:
                    if col_name in col_indices:
                        value = row[col_indices[col_name]]
                        row_data.append(self._clean_cell_value(value))
                    else:
                        row_data.append('')
                
                # Skip empty rows
                if any(row_data[1:]):
                    data.append(row_data)
            
            workbook.close()
            self.logger.info(f"Read {len(data)} rows from Excel")
            self.logger.log_function_end("read_excel_with_clean_columns")
            return data
            
        except Exception as e:
            self.logger.log_exception(e, "Read Excel with clean columns")
            raise
    
    def parse_colored_cells(
        self,
        file_path: str,
        sheet_name: str,
        column_names: List[str],
        start_row: int,
        color_map: Dict[Any, str]
    ) -> List[List[Any]]:
        """
        Parse Excel cells with color mapping
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name
            column_names: List of column names
            start_row: Starting row
            color_map: Dictionary mapping colors to status
        
        Returns:
            List of row data with color status
        """
        try:
            self.logger.log_function_start(
                "parse_colored_cells",
                file=file_path,
                sheet=sheet_name
            )
            
            if self.excel_handler:
                # Use ExcelHandler's colored cell parsing
                if hasattr(self.excel_handler, 'parse_colored_cells'):
                    data = self.excel_handler.parse_colored_cells(
                        file_path, sheet_name, column_names, start_row, color_map
                    )
                    return data
            
            # Fallback implementation
            workbook = load_workbook(file_path)
            sheet = workbook[sheet_name]
            
            # Find header and column indices
            header_row = self._find_header_row(sheet, column_names)
            col_indices = self._get_column_indices(sheet, header_row, column_names)
            
            data = []
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1 + start_row),
                start=header_row + 1 + start_row
            ):
                row_data = [row_idx]
                
                for col_name in column_names:
                    if col_name in col_indices:
                        cell = row[col_indices[col_name]]
                        value = self._clean_cell_value(cell.value)
                        row_data.append(value)
                
                # Get color status
                first_data_cell = row[col_indices[column_names[0]]]
                color_status = self._get_color_status(first_data_cell, color_map)
                row_data.append(color_status)
                
                if any(row_data[1:-1]):
                    data.append(row_data)
            
            workbook.close()
            self.logger.log_function_end("parse_colored_cells")
            return data
            
        except Exception as e:
            self.logger.log_exception(e, "Parse colored cells")
            raise
    
    def create_report_excel(
        self,
        output_path: str,
        sheets_data: Dict[str, List[List[Any]]]
    ) -> bool:
        """
        Create Excel report with multiple sheets
        
        Args:
            output_path: Output file path
            sheets_data: Dictionary of sheet_name: data
        
        Returns:
            bool: True if successful
        """
        try:
            self.logger.log_function_start("create_report_excel", output=output_path)
            
            workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            for sheet_name, data in sheets_data.items():
                sheet = workbook.create_sheet(title=sheet_name)
                
                if data:
                    # Write headers
                    if isinstance(data[0], dict):
                        headers = list(data[0].keys())
                        sheet.append(headers)
                        
                        # Style headers
                        for cell in sheet[1]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color="366092",
                                end_color="366092",
                                fill_type="solid"
                            )
                            cell.font = Font(color="FFFFFF", bold=True)
                        
                        # Write data
                        for row_data in data:
                            sheet.append(list(row_data.values()))
                    else:
                        # Write list data
                        for row_data in data:
                            sheet.append(row_data)
                
                # Auto-adjust column widths
                self._auto_adjust_columns(sheet)
            
            workbook.save(output_path)
            workbook.close()
            
            self.logger.log_file_operation("create", output_path, "Success")
            self.logger.log_function_end("create_report_excel")
            return True
            
        except Exception as e:
            self.logger.log_exception(e, "Create report Excel")
            return False
    
    def get_column_number(self, file_path: str, sheet_name: str, column_name: str) -> Optional[int]:
        """
        Get column number by column name
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name
            column_name: Column name to find
        
        Returns:
            Column number (1-indexed) or None
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]
            
            for row in sheet.iter_rows(max_row=10, values_only=True):
                for idx, cell in enumerate(row, start=1):
                    if cell and column_name.lower() in str(cell).lower().strip():
                        workbook.close()
                        return idx
            
            workbook.close()
            return None
            
        except Exception as e:
            self.logger.log_exception(e, "Get column number")
            return None
    
    def _clean_cell_value(self, value: Any) -> str:
        """Clean cell value removing HTML and extra spaces"""
        if value is None:
            return ''
        
        value_str = str(value).strip()
        
        # Remove HTML tags if present
        if '<' in value_str and '>' in value_str:
            soup = BeautifulSoup(value_str, 'html.parser')
            value_str = soup.get_text()
        
        return value_str.strip()
    
    def _find_header_row(self, sheet, column_names: List[str]) -> int:
        """Find header row in sheet"""
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if any(col in str(cell).strip() if cell else '' 
                  for cell in row for col in column_names):
                return row_idx
        raise ValueError("Header row not found")
    
    def _get_column_indices(self, sheet, header_row: int, column_names: List[str]) -> Dict[str, int]:
        """Get column indices from header row"""
        header = list(sheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True
        ))[0]
        
        col_indices = {}
        for col_name in column_names:
            for idx, cell in enumerate(header):
                if cell and col_name.lower() in str(cell).lower().strip():
                    col_indices[col_name] = idx
                    break
        return col_indices
    
    def _get_color_status(self, cell, color_map: Dict) -> str:
        """Get color status from cell"""
        if hasattr(cell, 'fill') and cell.fill.start_color:
            color = cell.fill.start_color.rgb
            if color in color_map:
                return color_map[color]
        return "Unknown"
    
    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
